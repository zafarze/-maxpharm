"""
Парсер Excel-файла со списком бонусов докторов.

Ожидаемые колонки (порядок не важен, заголовок ищется по слову «Клиент»):
    Дата | Организация | Менеджер | Область | Регион | Объект |
    Группа | Клиент | Телефон | Специальность | Утв СДК
"""
import datetime
import openpyxl

# Заголовок Excel → ключ внутреннего словаря
HEADER_ALIASES = {
    'дата': 'row_date',
    'организация': 'organization',
    'менеджер': 'manager',
    'область': 'oblast',
    'регион': 'region',
    'объект': 'object_name',
    'группа': 'group_name',
    'клиент': 'client_name',
    'телефон': 'phone',
    'специальность': 'specialty',
    'утв сдк': 'amount',
    'утв.сдк': 'amount',
    'сумма': 'amount',
}


def normalize_phone(raw):
    """Возвращает последние 9 цифр номера (TJ/UZ-формат) либо None."""
    if raw is None:
        return None
    digits = ''.join(ch for ch in str(raw) if ch.isdigit())
    if not digits:
        return None
    return digits[-9:]


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_date(v):
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime.combine(v, datetime.time())
    if v is None or v == '':
        return None
    for fmt in ('%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(str(v).strip(), fmt)
        except ValueError:
            continue
    return None


def _to_float(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_header(ws, scan_rows=10):
    """Ищет строку, содержащую ячейку «Клиент». Возвращает (row_idx, {key: col_idx})."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1):
        normalized = {}
        has_client = False
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            if key in HEADER_ALIASES:
                normalized[HEADER_ALIASES[key]] = col_idx
                if HEADER_ALIASES[key] == 'client_name':
                    has_client = True
        if has_client:
            return row_idx, normalized
    return None, {}


def parse_excel(file_path):
    """
    Возвращает (rows, errors).
      rows  — список словарей с ключами из HEADER_ALIASES + 'phone' (нормализованный).
      errors — список текстов ошибок (битые суммы, неизвестные группы и т.п.).
    Поднимает ValueError, если шапка не найдена или файл нечитаем.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось открыть файл: {e}")

    ws = wb.active
    header_row, header_map = _find_header(ws)
    if not header_row:
        raise ValueError("Не найден заголовок таблицы (колонка «Клиент»).")
    if 'amount' not in header_map:
        raise ValueError("Не найдена колонка с суммой («Утв СДК»).")
    if 'group_name' not in header_map:
        raise ValueError("Не найдена колонка «Группа».")

    rows = []
    errors = []

    def cell(row, key):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1
    ):
        if not row or not any(c is not None and str(c).strip() != '' for c in row):
            continue

        client_name = _str(cell(row, 'client_name'))
        if not client_name:
            continue

        amount = _to_float(cell(row, 'amount'))
        group = _str(cell(row, 'group_name')) or ''
        phone_raw = cell(row, 'phone')

        rows.append({
            'row_date': _to_date(cell(row, 'row_date')),
            'organization': _str(cell(row, 'organization')),
            'manager': _str(cell(row, 'manager')),
            'oblast': _str(cell(row, 'oblast')),
            'region': _str(cell(row, 'region')),
            'object_name': _str(cell(row, 'object_name')),
            'group_name': group,
            'client_name': client_name,
            'phone': normalize_phone(phone_raw),
            'specialty': _str(cell(row, 'specialty')),
            'amount': amount,
        })

    wb.close()
    return rows, errors
