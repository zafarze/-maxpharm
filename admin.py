import html
import os
import io
import tempfile
import datetime
import time
import threading

from openpyxl import Workbook
from telebot.types import (
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton
)
from telebot.apihelper import ApiTelegramException
from sqlalchemy import func
from database import get_session
from models import Doctor, BonusAck, FeedbackMessage, BroadcastHistory, Survey, SurveyQuestion, SurveyResponse, SurveyAnswer
from translations import get_text, TRANSLATIONS
from excel_parser import parse_excel
from bonus_service import process_upload
import config

# In-memory state
_broadcast_pending = {}     # {chat_id: (text, lang)}
_admin_lang = {}            # {chat_id: lang_code}
_broadcast_running = {}     # {admin_id: {'cancel': False, 'started_at': ...}}
_broadcast_running_lock = threading.Lock()

_survey_pending = {}        # {chat_id: {'questions': [str], 'mode': 'manual'|'excel', 'lang': str}}
_survey_running = {}        # {admin_id: {'cancel': False, 'survey_id': int|None, 'started_at': dt}}
_survey_lock = threading.Lock()

MAX_SURVEY_QUESTIONS = 50


def is_admin(user_id):
    """Return True if the given Telegram user ID matches the configured admin."""
    return str(user_id) in config.ADMIN_IDS


def get_admin_lang(chat_id):
    return _admin_lang.get(int(chat_id), 'ru')


def _set_admin_lang(chat_id, lang):
    _admin_lang[int(chat_id)] = lang


def get_admin_menu_keyboard(lang):
    kb = ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    # Row 1 — Excel upload (full width, primary action)
    kb.add(KeyboardButton(get_text(lang, 'admin_upload_excel')))
    # Row 2 — Survey composer + survey results
    kb.add(
        KeyboardButton(get_text(lang, 'admin_survey')),
        KeyboardButton(get_text(lang, 'admin_survey_results')),
    )
    # Row 3 — Stats + activity
    kb.add(
        KeyboardButton(get_text(lang, 'admin_stats')),
        KeyboardButton(get_text(lang, 'admin_engagement')),
    )
    # Row 4 — Doctors export + language
    kb.add(
        KeyboardButton(get_text(lang, 'admin_doctors')),
        KeyboardButton(get_text(lang, 'change_lang')),
    )
    return kb


def _matches(text, key):
    """Check if text matches the translation of key in any language."""
    return text in [TRANSLATIONS[l].get(key, '') for l in TRANSLATIONS]


def handle_admin_message(bot, message):
    """Route an admin text message to the correct feature handler."""
    lang = get_admin_lang(message.chat.id)
    text = message.text

    if _matches(text, 'admin_stats'):
        _send_stats(bot, message.chat.id, lang)
    elif _matches(text, 'admin_doctors'):
        _export_doctors_excel(bot, message.chat.id, lang)
    elif _matches(text, 'admin_upload_excel'):
        _start_excel_upload(bot, message, lang)
    elif _matches(text, 'admin_engagement'):
        _export_engagement_excel(bot, message.chat.id, lang)
    elif _matches(text, 'admin_survey_results'):
        _show_survey_picker(bot, message.chat.id, lang)
    elif _matches(text, 'admin_survey'):
        _start_survey(bot, message, lang)
    elif _matches(text, 'change_lang'):
        _show_lang_selector(bot, message.chat.id, lang)
    else:
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_menu_hint'),
            reply_markup=get_admin_menu_keyboard(lang)
        )


# ─── LANGUAGE CHANGE ─────────────────────────────────────────────────────────

def _show_lang_selector(bot, chat_id, lang):
    kb = InlineKeyboardMarkup(row_width=3)
    kb.add(
        InlineKeyboardButton('🇷🇺 Русский', callback_data='asetlang_ru'),
        InlineKeyboardButton('🇹🇯 Тоҷикӣ', callback_data='asetlang_tj'),
        InlineKeyboardButton('🇬🇧 English', callback_data='asetlang_en')
    )
    bot.send_message(chat_id, get_text(lang, 'choose_lang'), reply_markup=kb)


# ─── STATISTICS ──────────────────────────────────────────────────────────────

def _send_stats(bot, chat_id, lang):
    session = get_session()
    try:
        total = session.query(Doctor).count()
        linked = session.query(Doctor).filter(
            ~Doctor.telegram_id.like('PENDING%')
        ).count()
        last_sync = session.query(func.max(Doctor.last_update)).scalar()
        last_sync_str = last_sync.strftime("%d.%m.%Y %H:%M") if last_sync else "—"

        text = get_text(
            lang, 'admin_stats_msg',
            total=total, linked=linked, unlinked=total - linked,
            last_sync=last_sync_str
        )
        bot.send_message(chat_id, text, parse_mode='HTML')
    finally:
        session.close()


# ─── BROADCAST ───────────────────────────────────────────────────────────────

def _send_one_with_retry(bot, doctor, text):
    """Send broadcast message to a single doctor; retry once on 429."""
    for attempt in range(2):
        try:
            bot.send_message(
                int(doctor.telegram_id),
                get_text(doctor.language or 'ru', 'admin_broadcast_notify', text=text),
                parse_mode='HTML'
            )
            return True
        except ApiTelegramException as e:
            if e.error_code == 429 and attempt == 0:
                retry_after = (e.result_json or {}).get('parameters', {}).get('retry_after', 1)
                time.sleep(min(retry_after + 1, 30))
                continue
            return False
        except Exception:
            return False
    return False


def _broadcast_thread(bot, admin_id, chat_id, text, lang, progress_msg_id):
    """Background thread that performs the actual broadcast."""
    session = get_session()
    bh_id = None
    success = 0
    failed = 0
    cancelled = False
    try:
        doctors = session.query(Doctor).filter(
            ~Doctor.telegram_id.like('PENDING%')
        ).all()
        total = len(doctors)

        bh = BroadcastHistory(
            sent_by=str(admin_id),
            text=text,
            target_count=total,
            status='running',
        )
        session.add(bh)
        session.commit()
        bh_id = bh.id

        # Update progress message immediately with real total (HIGH #6)
        stop_kb = InlineKeyboardMarkup()
        stop_kb.add(InlineKeyboardButton(
            get_text(lang, 'admin_broadcast_stop_btn'),
            callback_data='bc_stop'
        ))
        try:
            bot.edit_message_text(
                get_text(lang, 'admin_broadcast_progress',
                         sent=0, total=total, success=0, failed=0),
                chat_id, progress_msg_id,
                reply_markup=stop_kb,
            )
        except Exception:
            pass

        progress_step = max(50, total // 20)

        for idx, doc in enumerate(doctors, start=1):
            with _broadcast_running_lock:
                if _broadcast_running.get(admin_id, {}).get('cancel'):
                    cancelled = True
                    break

            if _send_one_with_retry(bot, doc, text):
                success += 1
            else:
                failed += 1

            # Update progress message every N sends
            if idx % progress_step == 0 or idx == total:
                try:
                    stop_kb = InlineKeyboardMarkup()
                    stop_kb.add(InlineKeyboardButton(
                        get_text(lang, 'admin_broadcast_stop_btn'),
                        callback_data='bc_stop'
                    ))
                    bot.edit_message_text(
                        get_text(lang, 'admin_broadcast_progress',
                                 sent=idx, total=total,
                                 success=success, failed=failed),
                        chat_id, progress_msg_id,
                        reply_markup=stop_kb
                    )
                except Exception:
                    pass

            time.sleep(0.05)

        # Finalize DB record
        if bh_id is not None:
            bh = session.query(BroadcastHistory).get(bh_id)
            if bh is not None:
                bh.success_count = success
                bh.failed_count = failed
                bh.status = 'cancelled' if cancelled else 'completed'
                bh.finished_at = datetime.datetime.utcnow()
                session.commit()

        # Final progress message — remove stop button
        final_key = 'admin_broadcast_cancelled_mid' if cancelled else 'admin_broadcast_final'
        try:
            bot.edit_message_text(
                get_text(lang, final_key, success=success, failed=failed),
                chat_id, progress_msg_id,
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup()
            )
        except Exception:
            pass

    except Exception as e:
        print(f"[broadcast_thread] crashed: {type(e).__name__}: {e}")
        # Mark DB record as failed so no orphaned 'running' rows remain (HIGH #4)
        if bh_id is not None:
            try:
                s = get_session()
                try:
                    bh = s.get(BroadcastHistory, bh_id)
                    if bh and bh.status == 'running':
                        bh.status = 'failed'
                        bh.finished_at = datetime.datetime.utcnow()
                        s.commit()
                finally:
                    s.close()
            except Exception:
                pass
        try:
            bot.edit_message_text(
                get_text(lang, 'admin_broadcast_final', success=success, failed=failed),
                chat_id, progress_msg_id,
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup()
            )
        except Exception:
            pass
    finally:
        session.close()
        with _broadcast_running_lock:
            _broadcast_running.pop(admin_id, None)


def _start_broadcast(bot, message, lang):
    kb = ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    kb.add(KeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn')))
    msg = bot.send_message(
        message.chat.id,
        get_text(lang, 'admin_broadcast_prompt'),
        parse_mode='HTML',
        reply_markup=kb
    )
    bot.register_next_step_handler(msg, lambda m: _process_broadcast_input(bot, m, lang))


def _start_broadcast_cancel_kb(lang):
    kb = ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    kb.add(KeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn')))
    return kb


def _process_broadcast_input(bot, message, lang):
    cancel_labels = [TRANSLATIONS[l].get('admin_broadcast_cancel_btn', '') for l in TRANSLATIONS]

    if not message.text:
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_broadcast_not_text'),
            reply_markup=_start_broadcast_cancel_kb(lang),
        )
        bot.register_next_step_handler(message, lambda m: _process_broadcast_input(bot, m, lang))
        return

    if message.text in cancel_labels:
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_cancelled'),
            reply_markup=get_admin_menu_keyboard(lang)
        )
        return

    if len(message.text) > 4000:
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_broadcast_too_long', length=len(message.text)),
            reply_markup=get_admin_menu_keyboard(lang)
        )
        return

    # Pre-flight HTML validation — catches unclosed tags before sending to 1500 recipients.
    def _is_html_well_formed(s: str) -> bool:
        from html.parser import HTMLParser

        class _Stack(HTMLParser):
            def __init__(self):
                super().__init__()
                self.stack = []
                self.ok = True

            def handle_starttag(self, tag, attrs):
                # void tags are self-closing
                if tag in {'br', 'hr', 'img'}:
                    return
                self.stack.append(tag)

            def handle_endtag(self, tag):
                if tag in {'br', 'hr', 'img'}:
                    return
                if not self.stack or self.stack[-1] != tag:
                    self.ok = False
                else:
                    self.stack.pop()

        p = _Stack()
        try:
            p.feed(s)
        except Exception:
            return False
        return p.ok and not p.stack

    if not _is_html_well_formed(message.text):
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_broadcast_html_error'),
            reply_markup=get_admin_menu_keyboard(lang),
        )
        return

    with _broadcast_running_lock:
        _broadcast_pending[message.chat.id] = (message.text, lang)

    session = get_session()
    try:
        count = session.query(Doctor).filter(
            ~Doctor.telegram_id.like('PENDING%')
        ).count()
    finally:
        session.close()

    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton(get_text(lang, 'admin_broadcast_send_btn'), callback_data='bc_confirm'),
        InlineKeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn'), callback_data='bc_cancel')
    )
    # Escape preview text only; the raw text is kept in _broadcast_pending for delivery
    preview_text = html.escape(message.text)
    bot.send_message(
        message.chat.id,
        get_text(lang, 'admin_broadcast_preview', text=preview_text, count=count),
        parse_mode='HTML',
        reply_markup=kb
    )
    bot.send_message(
        message.chat.id,
        get_text(lang, 'admin_broadcast_hint'),
        reply_markup=get_admin_menu_keyboard(lang)
    )


# ─── DOCTORS MANAGEMENT ──────────────────────────────────────────────────────

def _export_doctors_excel(bot, chat_id, lang):
    """Export all doctors as an .xlsx file. Used instead of paginated list
    because the doctor base is too large (1500+ rows)."""
    try:
        bot.send_chat_action(chat_id, 'upload_document')
    except Exception:
        pass

    session = get_session()
    try:
        doctors = session.query(Doctor).order_by(Doctor.full_name.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Doctors'
        ws.append([
            '#', 'ФИО', 'Телефон', 'Специальность', 'Адрес', 'Код',
            'В Telegram-боте', 'Telegram ID',
            'Текущий бонус', 'Общий баланс', 'Последнее обновление',
        ])
        for i, d in enumerate(doctors, start=1):
            registered = bool(d.telegram_id and not d.telegram_id.startswith('PENDING'))
            ws.append([
                i,
                d.full_name or '',
                d.phone or '',
                d.specialty or '',
                d.address or '',
                d.doctor_code or '',
                'Да' if registered else 'Нет',
                d.telegram_id if registered else '',
                float(d.monthly_bonus or 0.0),
                float(d.current_balance or 0.0),
                d.last_update.strftime('%d.%m.%Y %H:%M') if d.last_update else '',
            ])

        widths = {'A': 5, 'B': 35, 'C': 16, 'D': 22, 'E': 28, 'F': 12,
                  'G': 16, 'H': 18, 'I': 14, 'J': 14, 'K': 18}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"doctors_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        buf.name = filename

        bot.send_document(
            chat_id,
            buf,
            visible_file_name=filename,
            caption=get_text(lang, 'admin_doctors_export_caption', total=len(doctors)),
            parse_mode='HTML',
            reply_markup=get_admin_menu_keyboard(lang),
        )
    except Exception as e:
        print(f"[doctors_export] error for {chat_id}: {e}")
        bot.send_message(
            chat_id,
            get_text(lang, 'admin_doctors_export_error', error=html.escape(str(e))),
            reply_markup=get_admin_menu_keyboard(lang),
        )
    finally:
        session.close()


def _export_engagement_excel(bot, chat_id, lang):
    """Export doctor engagement (acks + feedback) as a two-sheet .xlsx file."""
    try:
        bot.send_chat_action(chat_id, 'upload_document')
    except Exception:
        pass

    session = get_session()
    try:
        wb = Workbook()

        # Sheet 1 — activity per registered doctor (non-PENDING only)
        ws1 = wb.active
        ws1.title = 'Активность'
        ws1.append([
            '#', 'ФИО', 'Телефон', 'Спасибо нажал', 'Раз нажимал',
            'Последний раз', 'Обратная связь',
        ])

        doctors = session.query(Doctor).filter(
            ~Doctor.telegram_id.like('PENDING%')
        ).order_by(Doctor.full_name.asc()).all()

        # Single aggregated query — no N+1.
        ack_stats = {
            row.doctor_ref_id: (row.cnt, row.last_at)
            for row in session.query(
                BonusAck.doctor_ref_id,
                func.count(BonusAck.id).label('cnt'),
                func.max(BonusAck.acked_at).label('last_at'),
            )
            .filter(BonusAck.doctor_ref_id.isnot(None))
            .group_by(BonusAck.doctor_ref_id).all()
        }

        # Pre-load all feedback (asc → chronological concat for sheet 1, reversed for sheet 2).
        all_feedbacks = session.query(FeedbackMessage).order_by(
            FeedbackMessage.sent_at.asc()
        ).all()
        fb_by_doctor = {}
        for fb in all_feedbacks:
            if fb.doctor_ref_id is not None:
                date_str = fb.sent_at.strftime('%d.%m.%Y') if fb.sent_at else '—'
                fb_by_doctor.setdefault(fb.doctor_ref_id, []).append(
                    f"[{date_str}] {fb.text or ''}"
                )

        for i, d in enumerate(doctors, start=1):
            count, last_at = ack_stats.get(d.id, (0, None))
            fb_text = ' | '.join(fb_by_doctor.get(d.id, []))
            ws1.append([
                i,
                d.full_name or '',
                d.phone or '',
                'Да' if count else 'Нет',
                count,
                last_at.strftime('%d.%m.%Y %H:%M') if last_at else '',
                fb_text,
            ])

        widths1 = {'A': 5, 'B': 35, 'C': 16, 'D': 14, 'E': 12, 'F': 18, 'G': 60}
        for col, w in widths1.items():
            ws1.column_dimensions[col].width = w

        # Sheet 2 — all feedback messages, newest first
        ws2 = wb.create_sheet('Обратная связь')
        ws2.append(['#', 'Дата', 'ФИО', 'Телефон', 'Telegram ID', 'Сообщение'])
        for i, f in enumerate(reversed(all_feedbacks), start=1):
            ws2.append([
                i,
                f.sent_at.strftime('%d.%m.%Y %H:%M') if f.sent_at else '',
                f.full_name or '',
                f.phone or '',
                f.telegram_id or '',
                f.text or '',
            ])

        widths2 = {'A': 5, 'B': 18, 'C': 30, 'D': 16, 'E': 14, 'F': 70}
        for col, w in widths2.items():
            ws2.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"engagement_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        buf.name = filename

        bot.send_document(
            chat_id,
            buf,
            visible_file_name=filename,
            caption=get_text(
                lang, 'admin_engagement_caption',
                doctors=len(doctors), feedbacks=len(all_feedbacks),
            ),
            parse_mode='HTML',
            reply_markup=get_admin_menu_keyboard(lang),
        )
    except Exception as e:
        print(f"[engagement_export] error for {chat_id}: {e}")
        bot.send_message(
            chat_id,
            get_text(lang, 'admin_engagement_error', error=html.escape(str(e))),
            reply_markup=get_admin_menu_keyboard(lang),
        )
    finally:
        session.close()


def _show_doctors_list(bot, chat_id, lang, search=None):
    session = get_session()
    try:
        query = session.query(Doctor)
        if search:
            query = query.filter(
                Doctor.full_name.ilike(f'%{search}%') |
                Doctor.phone.ilike(f'%{search}%')
            )
        total = query.count()
        doctors = query.order_by(Doctor.id.desc()).limit(15).all()

        if not doctors:
            key = 'admin_doctors_not_found' if search else 'admin_doctors_empty'
            bot.send_message(chat_id, get_text(lang, key))
            return

        if search:
            header = get_text(lang, 'admin_doctors_search_header', total=total, search=html.escape(search or ''))
        else:
            header = get_text(lang, 'admin_doctors_header', total=total)

        kb = InlineKeyboardMarkup(row_width=1)
        for doc in doctors:
            registered = doc.telegram_id and not doc.telegram_id.startswith('PENDING')
            icon = "✅" if registered else "⏳"
            label = f"{icon} {doc.full_name or 'Без имени'} | {doc.phone or '—'}"
            kb.add(InlineKeyboardButton(label, callback_data=f'dr_{doc.id}'))

        kb.add(InlineKeyboardButton(get_text(lang, 'admin_doctors_search_btn'), callback_data='dr_search'))
        bot.send_message(chat_id, header, parse_mode='HTML', reply_markup=kb)
    finally:
        session.close()


def _show_doctor_detail(bot, call, lang):
    doc_id = int(call.data[3:])
    session = get_session()
    try:
        doc = session.query(Doctor).filter_by(id=doc_id).first()
        if not doc:
            bot.answer_callback_query(call.id, "Not found")
            return

        registered = doc.telegram_id and not doc.telegram_id.startswith('PENDING')
        linked_key = 'admin_doctor_linked' if registered else 'admin_doctor_not_linked'
        last_upd = doc.last_update.strftime("%d.%m.%Y %H:%M") if doc.last_update else "—"

        text = get_text(
            lang, 'admin_doctor_detail',
            name=doc.full_name or '—',
            phone=doc.phone or '—',
            address=doc.address or '—',
            code=doc.doctor_code or '—',
            bonus=doc.monthly_bonus,
            balance=doc.current_balance,
            linked=get_text(lang, linked_key),
            telegram_id=doc.telegram_id,
            last_upd=last_upd
        )
        kb = InlineKeyboardMarkup().add(
            InlineKeyboardButton(get_text(lang, 'admin_doctors_back_btn'), callback_data='dr_back')
        )
        bot.edit_message_text(
            text, call.message.chat.id, call.message.message_id,
            parse_mode='HTML', reply_markup=kb
        )
    finally:
        session.close()


# ─── EXCEL UPLOAD ────────────────────────────────────────────────────────────

def _excel_cancel_kb(lang):
    kb = ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    kb.add(KeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn')))
    return kb


def _is_cancel(text):
    return text in [TRANSLATIONS[l].get('admin_broadcast_cancel_btn', '') for l in TRANSLATIONS]


def _start_excel_upload(bot, message, lang):
    msg = bot.send_message(
        message.chat.id,
        get_text(lang, 'admin_excel_prompt'),
        parse_mode='HTML',
        reply_markup=_excel_cancel_kb(lang)
    )
    bot.register_next_step_handler(msg, lambda m: _receive_excel(bot, m, lang))


def _receive_excel(bot, message, lang):
    if message.text and _is_cancel(message.text):
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_cancelled'),
            reply_markup=get_admin_menu_keyboard(lang)
        )
        return

    doc = getattr(message, 'document', None)
    if not doc or not (doc.file_name or '').lower().endswith(('.xlsx', '.xlsm')):
        msg = bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_excel_invalid'),
            reply_markup=_excel_cancel_kb(lang)
        )
        bot.register_next_step_handler(msg, lambda m: _receive_excel(bot, m, lang))
        return

    bot.send_message(message.chat.id, get_text(lang, 'admin_excel_processing'))

    tmp_path = None
    try:
        file_info = bot.get_file(doc.file_id)
        data = bot.download_file(file_info.file_path)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            f.write(data)
            tmp_path = f.name

        rows, errors = parse_excel(tmp_path)
        if not rows:
            bot.send_message(
                message.chat.id,
                get_text(lang, 'admin_excel_empty'),
                reply_markup=get_admin_menu_keyboard(lang)
            )
            return

        def _send(tg_id, text):
            try:
                # Язык получателя — для подписи кнопки.
                _s = get_session()
                try:
                    rcpt = _s.query(Doctor).filter_by(telegram_id=str(tg_id)).first()
                    rcpt_lang = (rcpt.language if rcpt and rcpt.language else 'ru')
                finally:
                    _s.close()
                kb = InlineKeyboardMarkup()
                kb.add(
                    InlineKeyboardButton(get_text(rcpt_lang, 'bonus_thanks_btn'), callback_data='bonus_thanks'),
                    InlineKeyboardButton(get_text(rcpt_lang, 'bonus_feedback_btn'), callback_data='bonus_feedback'),
                )
                bot.send_message(int(tg_id), text, parse_mode='HTML', reply_markup=kb)
                return True
            except Exception as e:
                print(f"[excel] send error to {tg_id}: {e}")
                return False

        stats = process_upload(
            rows=rows,
            file_name=doc.file_name,
            admin_id=message.chat.id,
            send_message_fn=_send,
        )
        bot.send_message(
            message.chat.id,
            get_text(
                lang, 'admin_excel_done',
                total=stats['total_rows'],
                clients=stats['unique_clients'],
                notified=stats['notified'],
                pending=stats['pending'],
                failed=stats['failed'],
            ),
            parse_mode='HTML',
            reply_markup=get_admin_menu_keyboard(lang)
        )
    except ValueError as e:
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_excel_error', error=html.escape(str(e))),
            reply_markup=get_admin_menu_keyboard(lang)
        )
    except Exception as e:
        print(f"[excel] unexpected: {e}")
        bot.send_message(
            message.chat.id,
            get_text(lang, 'admin_excel_error', error=html.escape(str(e))),
            reply_markup=get_admin_menu_keyboard(lang)
        )
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


# ─── SURVEY ───────────────────────────────────────────────────────────────────

def _start_survey(bot, message, lang):
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton(get_text(lang, 'admin_survey_mode_one'), callback_data='sv_mode_one'),
        InlineKeyboardButton(get_text(lang, 'admin_survey_mode_multi'), callback_data='sv_mode_multi'),
    )
    kb.add(
        InlineKeyboardButton(get_text(lang, 'admin_survey_mode_excel'), callback_data='sv_mode_excel'),
        InlineKeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn'), callback_data='sv_mode_cancel'),
    )
    bot.send_message(
        message.chat.id,
        get_text(lang, 'admin_survey_choose_mode'),
        parse_mode='HTML',
        reply_markup=kb,
    )


def _process_survey_one_question(bot, message, lang):
    cancel_labels = [TRANSLATIONS[l].get('admin_broadcast_cancel_btn', '') for l in TRANSLATIONS]
    if message.text in cancel_labels:
        bot.send_message(message.chat.id, get_text(lang, 'admin_cancelled'),
                         reply_markup=get_admin_menu_keyboard(lang))
        return

    if not message.text:
        bot.send_message(message.chat.id, get_text(lang, 'admin_broadcast_not_text'),
                         reply_markup=_start_broadcast_cancel_kb(lang))
        bot.register_next_step_handler(message,
                                       lambda m: _process_survey_one_question(bot, m, lang))
        return

    if len(message.text) > 3500:
        bot.send_message(message.chat.id,
                         get_text(lang, 'admin_broadcast_too_long', length=len(message.text)),
                         reply_markup=get_admin_menu_keyboard(lang))
        return

    with _survey_lock:
        _survey_pending[message.chat.id] = {
            'questions': [message.text],
            'mode': 'manual',
            'lang': lang,
        }

    _show_survey_preview(bot, message.chat.id, lang)


def _process_survey_multi_question(bot, message, lang):
    chat_id = message.chat.id
    cancel_labels = [TRANSLATIONS[l].get('admin_broadcast_cancel_btn', '') for l in TRANSLATIONS]
    done_labels = [TRANSLATIONS[l].get('admin_survey_done_btn', '') for l in TRANSLATIONS]

    if message.text in cancel_labels:
        with _survey_lock:
            _survey_pending.pop(chat_id, None)
        bot.send_message(chat_id, get_text(lang, 'admin_cancelled'),
                         reply_markup=get_admin_menu_keyboard(lang))
        return

    if not message.text:
        bot.send_message(chat_id, get_text(lang, 'admin_broadcast_not_text'),
                         reply_markup=_survey_multi_kb(lang))
        bot.register_next_step_handler(message,
                                       lambda m: _process_survey_multi_question(bot, m, lang))
        return

    if message.text in done_labels:
        with _survey_lock:
            pending = _survey_pending.get(chat_id)
        if not pending or not pending['questions']:
            bot.send_message(chat_id, get_text(lang, 'admin_survey_prompt_multi_first'),
                             reply_markup=_survey_multi_kb(lang))
            bot.register_next_step_handler(message,
                                           lambda m: _process_survey_multi_question(bot, m, lang))
            return
        _show_survey_preview(bot, chat_id, lang)
        return

    if len(message.text) > 3500:
        bot.send_message(chat_id,
                         get_text(lang, 'admin_broadcast_too_long', length=len(message.text)),
                         reply_markup=_survey_multi_kb(lang))
        bot.register_next_step_handler(message,
                                       lambda m: _process_survey_multi_question(bot, m, lang))
        return

    with _survey_lock:
        pending = _survey_pending.get(chat_id)
        if not pending:
            bot.send_message(chat_id, get_text(lang, 'admin_error'),
                             reply_markup=get_admin_menu_keyboard(lang))
            return
        pending['questions'].append(message.text)
        n = len(pending['questions'])

    if n > MAX_SURVEY_QUESTIONS:
        bot.send_message(chat_id,
                         get_text(lang, 'admin_survey_too_many', limit=MAX_SURVEY_QUESTIONS),
                         reply_markup=get_admin_menu_keyboard(lang))
        with _survey_lock:
            _survey_pending.pop(chat_id, None)
        return

    bot.send_message(chat_id, get_text(lang, 'admin_survey_question_added', n=n))
    bot.register_next_step_handler(message,
                                   lambda m: _process_survey_multi_question(bot, m, lang))


def _survey_multi_kb(lang):
    kb = ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    kb.add(
        KeyboardButton(get_text(lang, 'admin_survey_done_btn')),
        KeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn')),
    )
    return kb


def _process_survey_excel(bot, message, lang):
    if _is_cancel(message.text or ''):
        bot.send_message(message.chat.id, get_text(lang, 'admin_cancelled'),
                         reply_markup=get_admin_menu_keyboard(lang))
        return

    if not message.document:
        bot.send_message(message.chat.id, get_text(lang, 'admin_excel_invalid'),
                         reply_markup=_excel_cancel_kb(lang))
        bot.register_next_step_handler(message, lambda m: _process_survey_excel(bot, m, lang))
        return

    fname = (message.document.file_name or '').lower()
    if not fname.endswith('.xlsx'):
        bot.send_message(message.chat.id, get_text(lang, 'admin_excel_invalid'),
                         reply_markup=_excel_cancel_kb(lang))
        bot.register_next_step_handler(message, lambda m: _process_survey_excel(bot, m, lang))
        return

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded = bot.download_file(file_info.file_path)
        tmp.write(downloaded)
        tmp.close()

        questions = _parse_survey_excel(tmp.name)
    except Exception as e:
        bot.send_message(message.chat.id,
                         get_text(lang, 'admin_excel_error', error=html.escape(str(e))),
                         reply_markup=get_admin_menu_keyboard(lang))
        return
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass

    if not questions:
        bot.send_message(message.chat.id, get_text(lang, 'admin_survey_excel_empty'),
                         reply_markup=get_admin_menu_keyboard(lang))
        return

    with _survey_lock:
        _survey_pending[message.chat.id] = {
            'questions': questions,
            'mode': 'excel',
            'lang': lang,
        }

    _show_survey_preview(bot, message.chat.id, lang)


def _parse_survey_excel(file_path):
    """Reads .xlsx, returns list of question texts from column 'Вопрос' (or 'Question').

    Loads file bytes into memory first so the OS file handle is released immediately,
    which avoids Windows file-lock issues when the caller deletes the temp file.
    """
    import io as _io
    from openpyxl import load_workbook

    with open(file_path, 'rb') as fh:
        data = fh.read()

    wb = load_workbook(_io.BytesIO(data), data_only=True)
    try:
        ws = wb.active

        header_aliases = {'вопрос', 'question', 'вопросы', 'questions'}
        header_row_idx = None
        col_idx = None

        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            for c_idx, cell in enumerate(row):
                if cell is not None and str(cell).strip().lower() in header_aliases:
                    header_row_idx = r_idx
                    col_idx = c_idx
                    break
            if header_row_idx:
                break

        if header_row_idx is None:
            raise ValueError("Не найдена колонка 'Вопрос'.")

        questions = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if cell is None:
                continue
            s = str(cell).strip()
            if s:
                questions.append(s)

        if len(questions) > MAX_SURVEY_QUESTIONS:
            raise ValueError(
                f"Слишком много вопросов: {len(questions)}. Максимум {MAX_SURVEY_QUESTIONS}."
            )
    finally:
        wb.close()

    return questions


def _show_survey_preview(bot, chat_id, lang):
    with _survey_lock:
        pending = _survey_pending.get(chat_id)
    if not pending:
        bot.send_message(chat_id, get_text(lang, 'admin_error'),
                         reply_markup=get_admin_menu_keyboard(lang))
        return

    questions = pending['questions']
    escaped_qs = '\n'.join(
        f"{i}. {html.escape(q)}" for i, q in enumerate(questions, start=1)
    )

    session = get_session()
    try:
        target_count = session.query(Doctor).filter(
            ~Doctor.telegram_id.like('PENDING%')
        ).count()
    finally:
        session.close()

    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton(get_text(lang, 'admin_broadcast_send_btn'), callback_data='sv_confirm'),
        InlineKeyboardButton(get_text(lang, 'admin_broadcast_cancel_btn'), callback_data='sv_cancel'),
    )
    bot.send_message(
        chat_id,
        get_text(lang, 'admin_survey_preview',
                 count=len(questions), recipients=target_count, items=escaped_qs),
        parse_mode='HTML',
        reply_markup=kb,
    )
    bot.send_message(chat_id, get_text(lang, 'admin_broadcast_hint'),
                     reply_markup=get_admin_menu_keyboard(lang))


def _survey_thread(bot, admin_id, chat_id, lang, questions, progress_msg_id):
    """Background thread: persist Survey + questions, broadcast Q1 to all eligible doctors."""
    session = get_session()
    survey_id = None
    success = 0
    failed = 0
    cancelled = False

    try:
        survey = Survey(
            sent_by=str(admin_id),
            status='running',
            target_count=0,
        )
        session.add(survey)
        session.commit()
        survey_id = survey.id

        for i, qtext in enumerate(questions, start=1):
            session.add(SurveyQuestion(survey_id=survey_id, order=i, text=qtext))
        session.commit()

        with _survey_lock:
            if admin_id in _survey_running:
                _survey_running[admin_id]['survey_id'] = survey_id

        doctors = session.query(Doctor).filter(
            ~Doctor.telegram_id.like('PENDING%')
        ).all()
        total = len(doctors)
        survey.target_count = total
        session.commit()

        stop_kb = InlineKeyboardMarkup().add(
            InlineKeyboardButton(get_text(lang, 'admin_broadcast_stop_btn'),
                                 callback_data='sv_stop')
        )
        try:
            bot.edit_message_text(
                get_text(lang, 'admin_survey_progress',
                         sent=0, total=total, success=0, failed=0),
                chat_id, progress_msg_id, reply_markup=stop_kb,
            )
        except Exception:
            pass

        progress_step = max(50, total // 20)
        first_question_text = questions[0]
        escaped_first_q = html.escape(first_question_text)
        n_questions = len(questions)

        # Bulk-cancel old in-progress responses for all target doctors in one UPDATE
        doctor_ids = [d.id for d in doctors]
        session.query(SurveyResponse).filter(
            SurveyResponse.doctor_ref_id.in_(doctor_ids),
            SurveyResponse.status == 'in_progress',
        ).update({'status': 'cancelled'}, synchronize_session=False)
        session.commit()

        for idx, doc in enumerate(doctors, start=1):
            with _survey_lock:
                if _survey_running.get(admin_id, {}).get('cancel'):
                    cancelled = True
                    break

            response = SurveyResponse(
                survey_id=survey_id,
                doctor_ref_id=doc.id,
                telegram_id=doc.telegram_id,
                status='in_progress',
                current_question_idx=1,
            )
            session.add(response)
            session.commit()

            try:
                bot.send_message(
                    int(doc.telegram_id),
                    get_text(doc.language or 'ru', 'survey_question',
                             current=1, total=n_questions, text=escaped_first_q),
                    parse_mode='HTML',
                )
                success += 1
            except ApiTelegramException as e:
                if e.error_code == 429:
                    retry_after = (e.result_json or {}).get('parameters', {}).get('retry_after', 1)
                    time.sleep(min(retry_after + 1, 30))
                    try:
                        bot.send_message(
                            int(doc.telegram_id),
                            get_text(doc.language or 'ru', 'survey_question',
                                     current=1, total=n_questions, text=escaped_first_q),
                            parse_mode='HTML',
                        )
                        success += 1
                    except Exception:
                        failed += 1
                        response.status = 'cancelled'
                        session.commit()
                else:
                    failed += 1
                    response.status = 'cancelled'
                    session.commit()
            except Exception:
                failed += 1
                response.status = 'cancelled'
                session.commit()

            if idx % progress_step == 0 or idx == total:
                try:
                    bot.edit_message_text(
                        get_text(lang, 'admin_survey_progress',
                                 sent=idx, total=total,
                                 success=success, failed=failed),
                        chat_id, progress_msg_id, reply_markup=stop_kb,
                    )
                except Exception:
                    pass

            time.sleep(0.05)

        survey.success_count = success
        survey.failed_count = failed
        survey.status = 'cancelled' if cancelled else 'completed'
        survey.finished_at = datetime.datetime.utcnow()
        session.commit()

        final_key = 'admin_survey_cancelled_mid' if cancelled else 'admin_survey_final'
        try:
            bot.edit_message_text(
                get_text(lang, final_key,
                         sent=success + failed, total=total,
                         success=success, failed=failed),
                chat_id, progress_msg_id, reply_markup=InlineKeyboardMarkup(),
            )
        except Exception:
            pass

    except Exception as e:
        print(f"[survey_thread] crashed: {type(e).__name__}: {e}")
        if survey_id is not None:
            try:
                s2 = get_session()
                try:
                    sv = s2.get(Survey, survey_id)
                    if sv and sv.status == 'running':
                        sv.status = 'failed'
                        sv.finished_at = datetime.datetime.utcnow()
                        s2.commit()
                finally:
                    s2.close()
            except Exception:
                pass
    finally:
        session.close()
        with _survey_lock:
            _survey_running.pop(admin_id, None)


# ─── SURVEY RESULTS EXPORT ───────────────────────────────────────────────────

def _show_survey_picker(bot, chat_id, lang):
    session = get_session()
    try:
        surveys = session.query(Survey).order_by(Survey.id.desc()).limit(20).all()
        if not surveys:
            bot.send_message(chat_id, get_text(lang, 'admin_survey_results_empty'),
                             reply_markup=get_admin_menu_keyboard(lang))
            return

        kb = InlineKeyboardMarkup(row_width=1)
        for sv in surveys:
            sent_str = sv.sent_at.strftime('%d.%m.%Y %H:%M') if sv.sent_at else '—'
            label = f'#{sv.id} · {sent_str} · {sv.target_count} получателей · {sv.status}'
            kb.add(InlineKeyboardButton(label, callback_data=f'svres_{sv.id}'))

        bot.send_message(
            chat_id,
            get_text(lang, 'admin_survey_results_picker'),
            parse_mode='HTML',
            reply_markup=kb,
        )
    finally:
        session.close()


def _export_survey_results(bot, chat_id, lang, survey_id):
    try:
        bot.send_chat_action(chat_id, 'upload_document')
    except Exception:
        pass

    session = get_session()
    try:
        survey = session.get(Survey, survey_id)
        if not survey:
            bot.send_message(chat_id, get_text(lang, 'admin_survey_not_found'),
                             reply_markup=get_admin_menu_keyboard(lang))
            return

        questions = session.query(SurveyQuestion).filter_by(
            survey_id=survey_id
        ).order_by(SurveyQuestion.order.asc()).all()
        n = len(questions)

        responses = session.query(SurveyResponse).filter_by(
            survey_id=survey_id
        ).all()

        # Pre-load answers grouped by (response_id, question_id).
        answer_map = {}
        if responses:
            response_ids = [r.id for r in responses]
            for ans in session.query(SurveyAnswer).filter(
                SurveyAnswer.response_id.in_(response_ids)
            ).all():
                answer_map[(ans.response_id, ans.question_id)] = ans.text or ''

        # Pre-load doctor names/phones.
        doctor_ids = [r.doctor_ref_id for r in responses if r.doctor_ref_id]
        docs = {}
        if doctor_ids:
            docs = {d.id: d for d in session.query(Doctor).filter(Doctor.id.in_(doctor_ids)).all()}

        wb = Workbook()
        ws = wb.active
        ws.title = f'Опрос {survey_id}'

        # Row 1: column headers
        header = ['#', 'ФИО', 'Телефон', 'Статус'] + [f'Q{i}' for i in range(1, n + 1)]
        ws.append(header)
        # Row 2: actual question texts under Q-columns
        ws.append(['', '', '', 'Текст:'] + [q.text for q in questions])

        for i, r in enumerate(responses, start=1):
            doc = docs.get(r.doctor_ref_id) if r.doctor_ref_id else None
            name = (doc.full_name if doc else r.telegram_id) or ''
            phone = (doc.phone if doc else '') or ''
            if r.status == 'completed':
                status = 'Завершён'
            elif r.status == 'cancelled':
                status = 'Отменён'
            else:
                idx = r.current_question_idx or 1
                status = f'В процессе ({max(0, idx - 1)}/{n})'

            row = [i, name, phone, status]
            for q in questions:
                row.append(answer_map.get((r.id, q.id), ''))
            ws.append(row)

        # Column widths
        widths = {'A': 5, 'B': 30, 'C': 16, 'D': 22}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        from openpyxl.utils import get_column_letter
        for i in range(1, n + 1):
            ws.column_dimensions[get_column_letter(4 + i)].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"survey_{survey_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        buf.name = filename

        completed = sum(1 for r in responses if r.status == 'completed')
        bot.send_document(
            chat_id, buf, visible_file_name=filename,
            caption=get_text(lang, 'admin_survey_results_caption',
                             survey_id=survey_id,
                             total=len(responses), completed=completed,
                             questions=n),
            parse_mode='HTML',
            reply_markup=get_admin_menu_keyboard(lang),
        )
    except Exception as e:
        print(f"[survey_results_export] error: {type(e).__name__}: {e}")
        bot.send_message(chat_id,
                         get_text(lang, 'admin_survey_results_error',
                                  error=html.escape(str(e))),
                         reply_markup=get_admin_menu_keyboard(lang))
    finally:
        session.close()


# ─── CALLBACK HANDLER REGISTRATION ───────────────────────────────────────────

def register_admin_handlers(bot):
    """Register all admin callback query handlers into the provided bot instance."""

    @bot.callback_query_handler(
        func=lambda c: c.data.startswith('asetlang_') and is_admin(c.from_user.id)
    )
    def cb_set_lang(call):
        lang_code = call.data.split('_')[1]
        _set_admin_lang(call.message.chat.id, lang_code)
        bot.edit_message_text(
            get_text(lang_code, 'lang_changed'),
            call.message.chat.id, call.message.message_id
        )
        bot.send_message(
            call.message.chat.id,
            get_text(lang_code, 'admin_welcome'),
            parse_mode='HTML',
            reply_markup=get_admin_menu_keyboard(lang_code)
        )

    @bot.callback_query_handler(
        func=lambda c: c.data in ('bc_confirm', 'bc_cancel') and is_admin(c.from_user.id)
    )
    def cb_broadcast(call):
        admin_id = call.from_user.id
        chat_id = call.message.chat.id
        lang = get_admin_lang(chat_id)

        if call.data == 'bc_cancel':
            with _broadcast_running_lock:
                _broadcast_pending.pop(chat_id, None)
            bot.answer_callback_query(call.id)
            bot.edit_message_text(
                get_text(lang, 'admin_broadcast_cancelled'),
                chat_id, call.message.message_id,
                reply_markup=None,
            )
            return

        # bc_confirm — atomic check + claim (BLOCKER #3)
        with _broadcast_running_lock:
            if admin_id in _broadcast_running:
                bot.answer_callback_query(
                    call.id,
                    get_text(lang, 'admin_broadcast_already_running'),
                    show_alert=True,
                )
                return
            pending = _broadcast_pending.pop(chat_id, None)
            if not pending:
                bot.answer_callback_query(call.id)
                bot.send_message(chat_id, get_text(lang, 'admin_error'))
                return
            _broadcast_running[admin_id] = {
                'cancel': False,
                'started_at': datetime.datetime.utcnow(),
            }

        text, lang = pending
        bot.answer_callback_query(call.id)
        # Edit the preview message to initial progress state
        stop_kb = InlineKeyboardMarkup()
        stop_kb.add(InlineKeyboardButton(
            get_text(lang, 'admin_broadcast_stop_btn'),
            callback_data='bc_stop'
        ))
        try:
            bot.edit_message_text(
                get_text(lang, 'admin_broadcast_progress',
                         sent=0, total='?', success=0, failed=0),
                chat_id, call.message.message_id,
                reply_markup=stop_kb
            )
            progress_msg_id = call.message.message_id
        except Exception:
            # Fallback: send a new progress message
            prog_msg = bot.send_message(
                chat_id,
                get_text(lang, 'admin_broadcast_progress',
                         sent=0, total='?', success=0, failed=0),
                reply_markup=stop_kb
            )
            progress_msg_id = prog_msg.message_id

        threading.Thread(
            target=_broadcast_thread,
            args=(bot, admin_id, chat_id, text, lang, progress_msg_id),
            daemon=True
        ).start()

    @bot.callback_query_handler(
        func=lambda c: c.data == 'bc_stop' and is_admin(c.from_user.id)
    )
    def cb_broadcast_stop(call):
        admin_id = call.from_user.id
        lang = get_admin_lang(call.message.chat.id)
        with _broadcast_running_lock:
            if admin_id in _broadcast_running:
                _broadcast_running[admin_id]['cancel'] = True
                bot.answer_callback_query(call.id, 'Останавливаю...')
            else:
                bot.answer_callback_query(call.id)

    @bot.callback_query_handler(
        func=lambda c: c.data == 'dr_search' and is_admin(c.from_user.id)
    )
    def cb_doctor_search(call):
        lang = get_admin_lang(call.message.chat.id)
        msg = bot.send_message(
            call.message.chat.id,
            get_text(lang, 'admin_doctors_search_prompt')
        )
        bot.register_next_step_handler(
            msg, lambda m: _show_doctors_list(bot, m.chat.id, get_admin_lang(m.chat.id), search=m.text)
        )

    @bot.callback_query_handler(
        func=lambda c: (
            c.data.startswith('dr_')
            and c.data not in ('dr_search', 'dr_back')
            and is_admin(c.from_user.id)
        )
    )
    def cb_doctor_detail(call):
        lang = get_admin_lang(call.message.chat.id)
        _show_doctor_detail(bot, call, lang)

    @bot.callback_query_handler(
        func=lambda c: c.data == 'dr_back' and is_admin(c.from_user.id)
    )
    def cb_doctor_back(call):
        lang = get_admin_lang(call.message.chat.id)
        bot.delete_message(call.message.chat.id, call.message.message_id)
        _show_doctors_list(bot, call.message.chat.id, lang)

    @bot.callback_query_handler(
        func=lambda c: c.data.startswith('sv_mode_') and is_admin(c.from_user.id)
    )
    def cb_survey_mode(call):
        chat_id = call.message.chat.id
        lang = get_admin_lang(chat_id)
        mode = call.data[len('sv_mode_'):]
        bot.answer_callback_query(call.id)

        if mode == 'cancel':
            bot.edit_message_text(get_text(lang, 'admin_cancelled'),
                                  chat_id, call.message.message_id, reply_markup=None)
            return

        if mode == 'one':
            bot.edit_message_text(
                get_text(lang, 'admin_survey_prompt_one'),
                chat_id, call.message.message_id, reply_markup=None,
            )
            bot.send_message(chat_id, get_text(lang, 'admin_survey_hint'),
                             reply_markup=_start_broadcast_cancel_kb(lang))
            bot.register_next_step_handler_by_chat_id(
                chat_id, lambda m: _process_survey_one_question(bot, m, lang)
            )
        elif mode == 'multi':
            with _survey_lock:
                _survey_pending[chat_id] = {'questions': [], 'mode': 'manual', 'lang': lang}
            bot.edit_message_text(
                get_text(lang, 'admin_survey_prompt_multi_first'),
                chat_id, call.message.message_id, reply_markup=None,
            )
            bot.send_message(chat_id, get_text(lang, 'admin_survey_hint'),
                             reply_markup=_survey_multi_kb(lang))
            bot.register_next_step_handler_by_chat_id(
                chat_id, lambda m: _process_survey_multi_question(bot, m, lang)
            )
        elif mode == 'excel':
            bot.edit_message_text(
                get_text(lang, 'admin_survey_prompt_excel'),
                chat_id, call.message.message_id, reply_markup=None,
            )
            bot.send_message(chat_id, get_text(lang, 'admin_survey_hint'),
                             reply_markup=_excel_cancel_kb(lang))
            bot.register_next_step_handler_by_chat_id(
                chat_id, lambda m: _process_survey_excel(bot, m, lang)
            )

    @bot.callback_query_handler(
        func=lambda c: c.data in ('sv_confirm', 'sv_cancel') and is_admin(c.from_user.id)
    )
    def cb_survey(call):
        admin_id = call.from_user.id
        chat_id = call.message.chat.id
        lang = get_admin_lang(chat_id)

        if call.data == 'sv_cancel':
            with _survey_lock:
                _survey_pending.pop(chat_id, None)
            bot.answer_callback_query(call.id)
            bot.edit_message_text(get_text(lang, 'admin_cancelled'),
                                  chat_id, call.message.message_id, reply_markup=None)
            return

        # sv_confirm
        with _survey_lock:
            if admin_id in _survey_running:
                bot.answer_callback_query(call.id,
                                          get_text(lang, 'admin_survey_already_running'),
                                          show_alert=True)
                return
            pending = _survey_pending.pop(chat_id, None)
            if not pending:
                bot.answer_callback_query(call.id)
                bot.send_message(chat_id, get_text(lang, 'admin_error'))
                return
            _survey_running[admin_id] = {'cancel': False, 'survey_id': None,
                                          'started_at': datetime.datetime.utcnow()}

        bot.answer_callback_query(call.id)
        questions = pending['questions']
        msg_lang = pending['lang']

        progress = bot.edit_message_text(
            get_text(msg_lang, 'admin_survey_progress',
                     sent=0, total='?', success=0, failed=0),
            chat_id, call.message.message_id,
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton(get_text(msg_lang, 'admin_broadcast_stop_btn'),
                                     callback_data='sv_stop')
            ),
        )

        threading.Thread(
            target=_survey_thread,
            args=(bot, admin_id, chat_id, msg_lang, questions, progress.message_id),
            daemon=True,
        ).start()

    @bot.callback_query_handler(
        func=lambda c: c.data == 'sv_stop' and is_admin(c.from_user.id)
    )
    def cb_survey_stop(call):
        admin_id = call.from_user.id
        lang = get_admin_lang(call.message.chat.id)
        with _survey_lock:
            if admin_id in _survey_running:
                _survey_running[admin_id]['cancel'] = True
        bot.answer_callback_query(call.id, get_text(lang, 'admin_broadcast_stopping'))

    @bot.callback_query_handler(
        func=lambda c: c.data.startswith('svres_') and is_admin(c.from_user.id)
    )
    def cb_survey_results(call):
        chat_id = call.message.chat.id
        lang = get_admin_lang(chat_id)
        try:
            survey_id = int(call.data[len('svres_'):])
        except (ValueError, TypeError):
            bot.answer_callback_query(call.id, 'Invalid survey id')
            return
        bot.answer_callback_query(call.id)
        _export_survey_results(bot, chat_id, lang, survey_id)
